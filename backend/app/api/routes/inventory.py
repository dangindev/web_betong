from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.application.inventory import (
    build_preclose_checklist,
    close_cost_period,
    create_cost_period,
    list_inventory_balances,
    open_cost_period,
    post_inventory_movement,
    post_inventory_stock_take,
    reopen_cost_period,
)
from app.core.dependencies import get_current_user, user_has_permission
from app.domain.models import User
from app.infrastructure.db import get_db

router = APIRouter(prefix="/api/v1", tags=["inventory"])


class InventoryMovementRequest(BaseModel):
    organization_id: str
    movement_type: str
    warehouse_id: str
    destination_warehouse_id: str | None = None
    material_id: str
    quantity: float | None = None
    quantity_delta: float | None = None
    unit_cost: float | None = None
    reference_no: str | None = None
    source_document_type: str | None = None
    source_document_id: str | None = None
    note: str | None = None
    transaction_at: datetime | None = None
    period_id: str | None = None


class InventoryStockTakeRequest(BaseModel):
    organization_id: str
    warehouse_id: str
    material_id: str
    counted_qty: float
    unit_cost: float | None = None
    note: str | None = None
    stock_take_date: date | None = None
    period_id: str | None = None


class CreateCostPeriodRequest(BaseModel):
    organization_id: str
    period_code: str
    start_date: date
    end_date: date
    note: str | None = None


class CostPeriodWorkflowRequest(BaseModel):
    organization_id: str
    note: str | None = None


def _has_any_permission(db: Session, user: User, modules: list[str], action: str) -> bool:
    for module in modules:
        if user_has_permission(db, user.id, module, action):
            return True
    if action == "read":
        for module in modules:
            if user_has_permission(db, user.id, module, "write"):
                return True
    return False


def _ensure_inventory_permission(db: Session, user: User, action: str) -> None:
    modules = [
        "warehouses",
        "inventory_ledger_entries",
        "inventory_stock_takes",
        "materials",
        "cost_periods",
        "inventory",
    ]
    if not _has_any_permission(db, user, modules, action):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Thiếu quyền thao tác kho")


def _ensure_costing_permission(db: Session, user: User, action: str) -> None:
    modules = ["cost_periods", "cost_centers", "cost_objects", "costing"]
    if not _has_any_permission(db, user, modules, action):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Thiếu quyền thao tác kỳ giá thành")


def _raise_service_error(exc: ValueError) -> None:
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


def _parse_rows(upload_file: UploadFile) -> list[dict[str, Any]]:
    suffix = upload_file.filename.rsplit(".", 1)[-1].lower() if upload_file.filename else ""
    content = upload_file.file.read()

    if suffix == "csv":
        decoded = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(decoded))
        return [dict(row) for row in reader]

    if suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        parsed_rows: list[dict[str, Any]] = []
        for row in rows[1:]:
            item: dict[str, Any] = {}
            for index, value in enumerate(row):
                if index < len(headers) and headers[index]:
                    item[headers[index]] = value
            parsed_rows.append(item)
        return parsed_rows

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chỉ hỗ trợ CSV/XLSX")


def _parse_optional_datetime(raw: Any) -> datetime | None:
    if raw in (None, ""):
        return None
    if isinstance(raw, datetime):
        return raw
    return datetime.fromisoformat(str(raw))


@router.post("/inventory/movements")
def create_inventory_movement(
    payload: InventoryMovementRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _ensure_inventory_permission(db, current_user, "write")
    try:
        return post_inventory_movement(
            db,
            organization_id=payload.organization_id,
            actor_user_id=current_user.id,
            movement_type=payload.movement_type,
            warehouse_id=payload.warehouse_id,
            destination_warehouse_id=payload.destination_warehouse_id,
            material_id=payload.material_id,
            quantity=payload.quantity,
            quantity_delta=payload.quantity_delta,
            unit_cost=payload.unit_cost,
            reference_no=payload.reference_no,
            source_document_type=payload.source_document_type,
            source_document_id=payload.source_document_id,
            note=payload.note,
            transaction_at=payload.transaction_at,
            period_id=payload.period_id,
        )
    except ValueError as exc:
        _raise_service_error(exc)


@router.post("/inventory/stock-takes")
def create_inventory_stock_take(
    payload: InventoryStockTakeRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _ensure_inventory_permission(db, current_user, "write")
    try:
        return post_inventory_stock_take(
            db,
            organization_id=payload.organization_id,
            actor_user_id=current_user.id,
            warehouse_id=payload.warehouse_id,
            material_id=payload.material_id,
            counted_qty=payload.counted_qty,
            unit_cost=payload.unit_cost,
            note=payload.note,
            stock_take_date=payload.stock_take_date,
            period_id=payload.period_id,
        )
    except ValueError as exc:
        _raise_service_error(exc)


@router.get("/inventory/balances")
def get_inventory_balances(
    organization_id: str = Query(...),
    warehouse_id: str | None = Query(default=None),
    material_id: str | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _ensure_inventory_permission(db, current_user, "read")
    return {
        "items": list_inventory_balances(
            db,
            organization_id=organization_id,
            warehouse_id=warehouse_id,
            material_id=material_id,
        )
    }


@router.post("/inventory/import-receipts")
def import_inventory_receipts(
    organization_id: str = Query(...),
    dry_run: bool = Query(default=False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _ensure_inventory_permission(db, current_user, "write")

    rows = _parse_rows(file)
    valid_rows = 0
    created = 0
    errors: list[dict[str, Any]] = []
    preview: list[dict[str, Any]] = []

    for index, row in enumerate(rows, start=2):
        try:
            warehouse_id = str(row.get("warehouse_id") or "").strip()
            material_id = str(row.get("material_id") or "").strip()
            quantity_raw = row.get("quantity")
            if not warehouse_id or not material_id or quantity_raw in (None, ""):
                raise ValueError("Thiếu warehouse_id/material_id/quantity")

            quantity = float(quantity_raw)
            payload = {
                "warehouse_id": warehouse_id,
                "material_id": material_id,
                "quantity": quantity,
                "unit_cost": float(row["unit_cost"]) if row.get("unit_cost") not in (None, "") else None,
                "reference_no": str(row.get("reference_no") or "").strip() or None,
                "source_document_type": str(row.get("source_document_type") or "").strip() or "import_receipt",
                "source_document_id": str(row.get("source_document_id") or "").strip() or None,
                "note": str(row.get("note") or "").strip() or None,
                "transaction_at": _parse_optional_datetime(row.get("transaction_at")),
                "period_id": str(row.get("period_id") or "").strip() or None,
            }

            valid_rows += 1
            if len(preview) < 20:
                preview.append(payload)

            if not dry_run:
                post_inventory_movement(
                    db,
                    organization_id=organization_id,
                    actor_user_id=current_user.id,
                    movement_type="receipt",
                    warehouse_id=payload["warehouse_id"],
                    destination_warehouse_id=None,
                    material_id=payload["material_id"],
                    quantity=payload["quantity"],
                    quantity_delta=None,
                    unit_cost=payload["unit_cost"],
                    reference_no=payload["reference_no"],
                    source_document_type=payload["source_document_type"],
                    source_document_id=payload["source_document_id"],
                    note=payload["note"],
                    transaction_at=payload["transaction_at"],
                    period_id=payload["period_id"],
                )
                created += 1

        except Exception as exc:  # noqa: BLE001
            errors.append(
                {
                    "row_number": index,
                    "error": str(exc),
                    "row": row,
                }
            )

    return {
        "resource": "inventory_receipts",
        "dry_run": dry_run,
        "total_rows": len(rows),
        "valid_rows": valid_rows,
        "invalid_rows": len(errors),
        "created": created,
        "skipped": len(errors),
        "preview": preview,
        "errors": errors,
    }


@router.get("/inventory/export-snapshot")
def export_inventory_snapshot(
    organization_id: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    _ensure_inventory_permission(db, current_user, "read")

    rows = list_inventory_balances(db, organization_id=organization_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ton_kho"
    sheet.append(
        [
            "warehouse_code",
            "warehouse_name",
            "material_code",
            "material_name",
            "available_qty",
            "last_transaction_at",
        ]
    )

    for row in rows:
        sheet.append(
            [
                row.get("warehouse_code"),
                row.get("warehouse_name"),
                row.get("material_code"),
                row.get("material_name"),
                row.get("available_qty"),
                row.get("last_transaction_at"),
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="inventory_snapshot.xlsx"'},
    )


@router.post("/costing/periods")
def create_period(
    payload: CreateCostPeriodRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _ensure_costing_permission(db, current_user, "write")
    try:
        return create_cost_period(
            db,
            organization_id=payload.organization_id,
            period_code=payload.period_code,
            start_date=payload.start_date,
            end_date=payload.end_date,
            note=payload.note,
        )
    except ValueError as exc:
        _raise_service_error(exc)


@router.post("/costing/periods/{period_id}/open")
def open_period(
    period_id: str,
    payload: CostPeriodWorkflowRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _ensure_costing_permission(db, current_user, "write")
    try:
        return open_cost_period(
            db,
            organization_id=payload.organization_id,
            period_id=period_id,
            actor_user_id=current_user.id,
            note=payload.note,
        )
    except ValueError as exc:
        _raise_service_error(exc)


@router.get("/costing/periods/{period_id}/preclose-checklist")
def get_preclose_checklist(
    period_id: str,
    organization_id: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _ensure_costing_permission(db, current_user, "read")
    try:
        return build_preclose_checklist(
            db,
            organization_id=organization_id,
            period_id=period_id,
        )
    except ValueError as exc:
        _raise_service_error(exc)


@router.post("/costing/periods/{period_id}/close")
def close_period(
    period_id: str,
    payload: CostPeriodWorkflowRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _ensure_costing_permission(db, current_user, "write")
    try:
        return close_cost_period(
            db,
            organization_id=payload.organization_id,
            period_id=period_id,
            actor_user_id=current_user.id,
            note=payload.note,
        )
    except ValueError as exc:
        _raise_service_error(exc)


@router.post("/costing/periods/{period_id}/reopen")
def reopen_period(
    period_id: str,
    payload: CostPeriodWorkflowRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _ensure_costing_permission(db, current_user, "write")
    try:
        return reopen_cost_period(
            db,
            organization_id=payload.organization_id,
            period_id=period_id,
            actor_user_id=current_user.id,
            note=payload.note,
        )
    except ValueError as exc:
        _raise_service_error(exc)
