from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.application.registry import (
    IMPORT_EXPORT_RESOURCES,
    MODEL_REGISTRY,
    allowed_columns,
    serialize_instance,
)
from app.core.config import settings
from app.core.dependencies import (
    apply_scope_filters,
    get_current_user,
    get_scope_context,
    payload_in_scope,
    row_in_scope,
    user_has_permission,
)
from app.domain.models import Attachment, User
from app.infrastructure.db import get_db
from app.infrastructure.geocode import geocode_adapter

router = APIRouter(prefix="/api/v1", tags=["integration"])

REQUIRED_FIELDS: dict[str, list[str]] = {
    "customers": ["organization_id", "code", "name"],
    "project_sites": ["organization_id", "customer_id", "code", "site_name", "address_line"],
    "vehicles": ["organization_id", "plate_no"],
    "materials": ["organization_id", "code", "name"],
    "price_rules": ["price_book_id", "rule_type", "rule_name"],
    "pour_requests": [
        "organization_id",
        "request_no",
        "customer_id",
        "site_id",
        "concrete_product_id",
        "requested_volume_m3",
    ],
}


@router.post("/geocode")
def geocode_address(
    payload: dict[str, str],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    if not user_has_permission(db, current_user.id, "project_sites", "read"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Missing permission")

    address = payload.get("address", "").strip()
    if not address:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="address is required")

    result = geocode_adapter.geocode(address)
    return {
        "address": result.address,
        "latitude": result.latitude,
        "longitude": result.longitude,
        "source": result.source,
    }



def _parse_rows(upload_file: UploadFile) -> list[dict[str, Any]]:
    suffix = upload_file.filename.rsplit(".", 1)[-1].lower() if upload_file.filename else ""
    content = upload_file.file.read()

    if suffix == "csv":
        decoded = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(decoded))
        return [dict(row) for row in reader]

    if suffix in {"xlsx", "xlsm"}:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        parsed: list[dict[str, Any]] = []
        for row in rows[1:]:
            item: dict[str, Any] = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    item[headers[idx]] = value
            parsed.append(item)
        return parsed

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only CSV/XLSX are supported")



def _sanitize_import_row(row: dict[str, Any], allowed: set[str]) -> dict[str, Any]:
    sanitized: dict[str, Any] = {}
    for key, value in row.items():
        normalized_key = str(key).strip()
        if normalized_key in allowed and normalized_key not in {"created_at", "updated_at"}:
            if isinstance(value, str):
                stripped = value.strip()
                sanitized[normalized_key] = stripped if stripped != "" else None
            else:
                sanitized[normalized_key] = value
    return sanitized



def _validate_required_fields(resource: str, data: dict[str, Any]) -> list[str]:
    required = REQUIRED_FIELDS.get(resource, [])
    return [field for field in required if data.get(field) in (None, "")]



def _build_import_result(
    resource: str,
    rows: list[dict[str, Any]],
    model: type,
    scope,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    allowed = allowed_columns(model)
    valid_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for index, row in enumerate(rows, start=2):
        data = _sanitize_import_row(row, allowed)
        if not data:
            errors.append(
                {
                    "row_number": index,
                    "error": "No importable columns",
                    "row": row,
                }
            )
            continue

        missing_fields = _validate_required_fields(resource, data)
        if missing_fields:
            errors.append(
                {
                    "row_number": index,
                    "error": f"Missing required fields: {', '.join(missing_fields)}",
                    "row": data,
                }
            )
            continue

        if not payload_in_scope(model, data, scope):
            errors.append(
                {
                    "row_number": index,
                    "error": "Row outside assigned scope",
                    "row": data,
                }
            )
            continue

        valid_rows.append(data)

    return valid_rows, errors


@router.post("/io/import/{resource}")
def import_resource(
    resource: str,
    dry_run: bool = Query(False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    if resource not in IMPORT_EXPORT_RESOURCES:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import not supported")

    if not user_has_permission(db, current_user.id, resource, "write"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Missing permission")

    model = MODEL_REGISTRY[resource]
    rows = _parse_rows(file)
    scope = get_scope_context(db, current_user.id)

    valid_rows, errors = _build_import_result(resource, rows, model, scope)

    created = 0
    if not dry_run:
        for data in valid_rows:
            db.add(model(**data))
            created += 1
        db.commit()

    return {
        "resource": resource,
        "dry_run": dry_run,
        "total_rows": len(rows),
        "valid_rows": len(valid_rows),
        "invalid_rows": len(errors),
        "created": created,
        "skipped": len(errors),
        "preview": valid_rows[:20],
        "errors": errors,
    }


@router.get("/io/export/{resource}")
def export_resource(
    resource: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    if resource not in IMPORT_EXPORT_RESOURCES:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Export not supported")

    if not user_has_permission(db, current_user.id, resource, "read"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Missing permission")

    model = MODEL_REGISTRY[resource]
    scope = get_scope_context(db, current_user.id)
    rows = db.execute(apply_scope_filters(model, select(model), scope)).scalars().all()
    serialized = [serialize_instance(row) for row in rows]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = resource

    if serialized:
        headers = list(serialized[0].keys())
        sheet.append(headers)
        for row in serialized:
            sheet.append([row.get(header) for header in headers])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{resource}.xlsx"'},
    )


@router.post("/attachments/upload")
def upload_attachment(
    entity_type: str = Form(...),
    entity_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    if not (
        user_has_permission(db, current_user.id, entity_type, "write")
        or user_has_permission(db, current_user.id, "attachments", "write")
    ):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Missing permission")

    entity_model = MODEL_REGISTRY.get(entity_type)
    if entity_model is not None:
        entity = db.get(entity_model, entity_id)
        if not entity:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entity not found")
        scope = get_scope_context(db, current_user.id)
        if not row_in_scope(entity_model, entity, scope):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Entity outside assigned scope")

    content = file.file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty file")

    safe_name = Path(file.filename or "upload.bin").name
    extension = Path(safe_name).suffix
    file_key = f"{entity_type}/{uuid4().hex}{extension}"

    upload_root = Path(settings.upload_dir)
    destination = upload_root / file_key
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(content)

    attachment = Attachment(
        entity_type=entity_type,
        entity_id=entity_id,
        file_key=file_key,
        file_name=safe_name,
        content_type=file.content_type,
        size_bytes=len(content),
        uploaded_by=current_user.id,
        uploaded_at=datetime.now(tz=timezone.utc),
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)

    return serialize_instance(attachment)
